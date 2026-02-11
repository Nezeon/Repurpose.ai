"""
Excel Report Generator
Generates multi-sheet Excel reports from drug repurposing search results.
"""

import io
from datetime import datetime
from typing import Any, Dict, Union

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.utils.logger import get_logger

logger = get_logger("excel_generator")

# Styling constants (matching platform design system)
HEADER_FILL = PatternFill(start_color="FFE600", end_color="FFE600", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="000000")
TITLE_FONT = Font(bold=True, size=16, color="000000")
SUBTITLE_FONT = Font(bold=True, size=12, color="333333")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

# Dimension-specific colors
SCI_FILL = PatternFill(start_color="00D4E8", end_color="00D4E8", fill_type="solid")
MKT_FILL = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
COMP_FILL = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
FEAS_FILL = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")


def _get(obj, key, default=None):
    """Get attribute or dict key — allows functions to accept both dict and Pydantic models."""
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def generate_excel_report(result: Union[dict, Any]) -> bytes:
    """
    Generate Excel report with multiple sheets.

    Accepts either a SearchResponse Pydantic model or a raw pipeline dict.

    Sheets:
    1. Summary — high-level overview and top opportunities
    2. Opportunities — all ranked indications with 4D scores
    3. Evidence — all evidence items from all agents
    4. Market Data — market and synthesis insights

    Returns: Excel file as bytes
    """
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, result)

    ws_opps = wb.create_sheet("Opportunities")
    _write_opportunities_sheet(ws_opps, result)

    ws_evidence = wb.create_sheet("Evidence")
    _write_evidence_sheet(ws_evidence, result)

    ws_market = wb.create_sheet("Market Data")
    _write_market_sheet(ws_market, result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _styled_header(ws, row, headers, fill=HEADER_FILL):
    """Write a styled header row."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, columns, min_width=10, max_width=60):
    """Auto-size column widths."""
    for col in range(1, columns + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max_len + 2


def _write_summary_sheet(ws, result):
    """Summary sheet with key metrics and top opportunities."""
    ws["A1"] = f"Drug Repurposing Report: {_get(result, 'drug_name', 'Unknown')}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    ws["A3"] = "Generated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Session ID"
    ws["B4"] = _get(result, 'session_id', 'N/A')
    ws["A5"] = "Execution Time"
    exec_time = _get(result, 'execution_time', 0) or 0
    ws["B5"] = f"{exec_time:.1f}s"

    for r in range(3, 6):
        ws.cell(row=r, column=1).font = Font(bold=True)

    # Key metrics
    ws["A7"] = "Key Metrics"
    ws["A7"].font = SUBTITLE_FONT
    ws["A8"] = "Total Opportunities"
    ranked = _get(result, 'ranked_indications', []) or []
    ws["B8"] = len(ranked)
    ws["A9"] = "Total Evidence Items"
    all_ev = _get(result, 'all_evidence', []) or []
    ws["B9"] = _get(result, 'total_evidence_count', len(all_ev))
    ws["A10"] = "Data Sources"
    agent_res = _get(result, 'agent_results', {}) or {}
    ws["B10"] = len(agent_res)

    for r in range(8, 11):
        ws.cell(row=r, column=1).font = Font(bold=True)

    # Top 5 opportunities table
    ws["A12"] = "Top 5 Opportunities"
    ws["A12"].font = SUBTITLE_FONT

    headers = ["Rank", "Indication", "Score", "Evidence Count", "Sources"]
    _styled_header(ws, 13, headers)

    indications = _get(result, 'enhanced_indications', []) or _get(result, 'ranked_indications', []) or []
    for idx, opp in enumerate(indications[:5], start=1):
        row = 13 + idx
        if isinstance(opp, dict):
            indication = opp.get("indication", "Unknown")
            cs = opp.get("composite_score", {})
            score = cs.get("overall_score", opp.get("confidence_score", 0)) if isinstance(cs, dict) else getattr(cs, "overall_score", 0)
            ev_count = opp.get("evidence_count", 0)
            sources = ", ".join(opp.get("supporting_sources", []))
        else:
            indication = getattr(opp, "indication", "Unknown")
            score = getattr(opp, "confidence_score", 0)
            ev_count = getattr(opp, "evidence_count", 0)
            sources = ", ".join(getattr(opp, "supporting_sources", []))

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=indication)
        ws.cell(row=row, column=3, value=round(score, 1))
        ws.cell(row=row, column=4, value=ev_count)
        ws.cell(row=row, column=5, value=sources)

        for c in range(1, 6):
            ws.cell(row=row, column=c).border = THIN_BORDER

    _auto_width(ws, 5)
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["E"].width = 40


def _write_opportunities_sheet(ws, result):
    """All ranked indications with 4D scores."""
    headers = [
        "Rank", "Indication", "Overall Score", "Confidence",
        "Scientific", "Market", "Competitive", "Feasibility",
        "Evidence Count", "Sources"
    ]
    _styled_header(ws, 1, headers)

    indications = _get(result, 'enhanced_indications', []) or _get(result, 'ranked_indications', []) or []
    for idx, opp in enumerate(indications, start=1):
        row = idx + 1
        if isinstance(opp, dict):
            indication = opp.get("indication", "Unknown")
            cs = opp.get("composite_score", {})
            if isinstance(cs, dict):
                overall = cs.get("overall_score", opp.get("confidence_score", 0))
                confidence = cs.get("confidence_level", "N/A")
                sci = cs.get("scientific_evidence", {}).get("score", 0) if isinstance(cs.get("scientific_evidence"), dict) else 0
                mkt = cs.get("market_opportunity", {}).get("score", 0) if isinstance(cs.get("market_opportunity"), dict) else 0
                comp = cs.get("competitive_landscape", {}).get("score", 0) if isinstance(cs.get("competitive_landscape"), dict) else 0
                feas = cs.get("development_feasibility", {}).get("score", 0) if isinstance(cs.get("development_feasibility"), dict) else 0
            else:
                overall = opp.get("confidence_score", 0)
                confidence = "N/A"
                sci = mkt = comp = feas = 0
            ev_count = opp.get("evidence_count", 0)
            sources = ", ".join(opp.get("supporting_sources", []))
        else:
            indication = getattr(opp, "indication", "Unknown")
            overall = getattr(opp, "confidence_score", 0)
            confidence = "N/A"
            sci = mkt = comp = feas = 0
            ev_count = getattr(opp, "evidence_count", 0)
            sources = ", ".join(getattr(opp, "supporting_sources", []))

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=indication)
        ws.cell(row=row, column=3, value=round(overall, 1))
        ws.cell(row=row, column=4, value=confidence)
        ws.cell(row=row, column=5, value=round(sci, 1))
        ws.cell(row=row, column=6, value=round(mkt, 1))
        ws.cell(row=row, column=7, value=round(comp, 1))
        ws.cell(row=row, column=8, value=round(feas, 1))
        ws.cell(row=row, column=9, value=ev_count)
        ws.cell(row=row, column=10, value=sources)

        for c in range(1, 11):
            ws.cell(row=row, column=c).border = THIN_BORDER

    _auto_width(ws, 10)
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["J"].width = 35


def _write_evidence_sheet(ws, result):
    """All evidence items across all indications."""
    headers = ["Source", "Indication", "Title", "Summary", "Date", "Relevance", "URL"]
    _styled_header(ws, 1, headers)

    row = 2
    seen = set()

    # Use enhanced_indications (which have evidence_items) or fall back to ranked_indications
    opps = _get(result, 'enhanced_indications', []) or _get(result, 'ranked_indications', []) or []
    for opp in opps:
        if isinstance(opp, dict):
            items = opp.get("evidence_items", []) or []
            opp_indication = opp.get("indication", "")
        else:
            items = getattr(opp, "evidence_items", []) or []
            opp_indication = getattr(opp, "indication", "")

        if not isinstance(items, list):
            items = []

        for ev in items:
            if isinstance(ev, dict):
                key = (ev.get("source", ""), ev.get("summary", "")[:50])
                if key in seen:
                    continue
                seen.add(key)
                ws.cell(row=row, column=1, value=ev.get("source", ""))
                ws.cell(row=row, column=2, value=ev.get("indication", opp_indication))
                ws.cell(row=row, column=3, value=ev.get("title", ""))
                ws.cell(row=row, column=4, value=ev.get("summary", ""))
                ws.cell(row=row, column=5, value=ev.get("date", ""))
                ws.cell(row=row, column=6, value=round(ev.get("relevance_score", 0) or 0, 2))
                ws.cell(row=row, column=7, value=ev.get("url", ""))
            else:
                key = (getattr(ev, "source", ""), getattr(ev, "summary", "")[:50])
                if key in seen:
                    continue
                seen.add(key)
                ws.cell(row=row, column=1, value=getattr(ev, "source", ""))
                ws.cell(row=row, column=2, value=getattr(ev, "indication", "") or opp_indication)
                ws.cell(row=row, column=3, value=getattr(ev, "title", ""))
                ws.cell(row=row, column=4, value=getattr(ev, "summary", ""))
                ws.cell(row=row, column=5, value=getattr(ev, "date", ""))
                ws.cell(row=row, column=6, value=round(getattr(ev, "relevance_score", 0) or 0, 2))
                ws.cell(row=row, column=7, value=getattr(ev, "url", ""))

            for c in range(1, 8):
                ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    _auto_width(ws, 7)
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["G"].width = 40


def _write_market_sheet(ws, result):
    """Market data and AI synthesis."""
    ws["A1"] = "Market & Strategic Insights"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # Enhanced opportunities market data
    enhanced_opps = _get(result, 'enhanced_opportunities', {}) or {}
    if enhanced_opps:
        ws["A3"] = "Market Segments by Indication"
        ws["A3"].font = SUBTITLE_FONT

        headers = ["Indication", "Segment", "Market Size", "CAGR", "Unmet Need", "Competition"]
        _styled_header(ws, 4, headers)

        row = 5
        for indication, data in enhanced_opps.items():
            if isinstance(data, dict):
                market = data.get("market_segment", {})
                if isinstance(market, dict):
                    ws.cell(row=row, column=1, value=indication)
                    ws.cell(row=row, column=2, value=market.get("segment_name", ""))
                    ws.cell(row=row, column=3, value=market.get("segment_size", ""))
                    ws.cell(row=row, column=4, value=market.get("growth_rate", ""))
                    ws.cell(row=row, column=5, value=market.get("unmet_need_level", ""))
                    ws.cell(row=row, column=6, value=market.get("competitive_intensity", ""))
                    for c in range(1, 7):
                        ws.cell(row=row, column=c).border = THIN_BORDER
                    row += 1

        _auto_width(ws, 6)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 35
        row += 1
    else:
        row = 3

    # AI Synthesis
    synthesis = _get(result, 'synthesis', '') or ''
    if synthesis:
        ws.cell(row=row, column=1, value="AI-Generated Strategic Summary")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        row += 1
        ws.cell(row=row, column=1, value=synthesis)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 150

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 30, 30)
